import playwright
from playwright.sync_api import Playwright, sync_playwright, expect
from openpyxl import load_workbook
from time import sleep
import os

acc_email = os.getenv('EMAIL')
acc_pwd = os.getenv('PWD')

wb = load_workbook('clients.xlsx')
sheet = wb['clients-data']


def run(playwright: Playwright) -> None:
    browser = playwright.firefox.launch(headless=False)
    context = browser.new_context()
    page = context.new_page()

    # login step
    page.goto("https://app.akaunting.com/auth/login")
    sleep(5)
    page.get_by_placeholder("Email").click()
    page.get_by_placeholder("Email").fill(acc_email)
    page.get_by_placeholder("Password").click()
    page.get_by_placeholder("Password").fill(acc_pwd)
    page.get_by_label("Remember Me").check()
    page.get_by_role("button", name="Login").click()
    sleep(3)
    page.goto("https://app.akaunting.com/252757")

    # clients section prepartion
    page.get_by_role("button", name="تأكيد").click()
    page.get_by_text("payments البيع expand_more").click()

    for i in range(2, sheet.max_row+1):

        name = sheet[f'A{i}'].value
        email = sheet[f'B{i}'].value
        phone = sheet[f'C{i}'].value
        tax_num = sheet[f'D{i}'].value
        address = sheet[f'E{i}'].value
        city = sheet[f'F{i}'].value
        zcode = sheet[f'G{i}'].value
        state = sheet[f'H{i}'].value
        country = sheet[f'I{i}'].value

        page.get_by_role("link", name="العميل").click()
        page.get_by_role("link", name="إضافة العملاء").click()

        sleep(2)

        page.get_by_placeholder("إدخال الاسم").click()
        page.get_by_placeholder("إدخال الاسم").fill(name)

        page.get_by_placeholder("إدخال البريد الإلكتروني").click()
        page.get_by_placeholder("إدخال البريد الإلكتروني").fill(email)

        page.get_by_placeholder("إدخال رقم الهاتف").click()
        page.get_by_placeholder("إدخال رقم الهاتف").fill(phone)

        page.get_by_placeholder("إدخال رقم الضريبة").click()
        page.get_by_placeholder("إدخال رقم الضريبة").fill(tax_num)

        page.get_by_placeholder("إدخال العنوان").click()
        page.get_by_placeholder("إدخال العنوان").fill(address)

        page.get_by_placeholder("إدخال Towns / Cities").click()
        page.get_by_placeholder("إدخال Towns / Cities").fill(city)

        page.get_by_placeholder("إدخال Postal / Zip Code").click()
        page.get_by_placeholder("إدخال Postal / Zip Code").fill(zcode)

        page.get_by_placeholder("إدخال Province / State").click()
        page.get_by_placeholder("إدخال Province / State").fill(state)

        page.get_by_role("button", name="حفظ").click()
        page.get_by_role("link", name="العميل").click()

    sleep(10)

    # ---------------------
    context.close()
    browser.close()


with sync_playwright() as playwright:
    run(playwright)
